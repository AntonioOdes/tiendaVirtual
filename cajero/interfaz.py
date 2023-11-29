
from tkinter import *
import os
from openpyxl import Workbook
from openpyxl import load_workbook
from CTkMessagebox import CTkMessagebox
from openpyxl.workbook import Workbook
import conexion


from customtkinter import *
from CTkListbox import *
from customtkinter import  CTkButton, CTkEntry, CTkImage, CTkLabel
import customtkinter as ctk
from PIL import ImageTk, Image



class Interfaz (object):
    
    def __init__(self) -> None:
        self.ventana=ctk.CTk()
        self.datos = conexion.Registro_de_datos()
        self.ventana.geometry("1240x720")
        self.grupo1 = ctk.CTkFrame(self.ventana, width= 440,height=600)
        self.grupo2 = ctk.CTkFrame(self.ventana, width= 440,height=600)
        altura = self.ventana.winfo_reqheight()
        anchura = self.ventana.winfo_reqwidth()
        altura_pantalla = self.ventana.winfo_screenheight()
        anchura_pantalla = self.ventana.winfo_screenwidth()
        #print(f"Altura: {altura}\nAnchura: {anchura}\nAltura de pantalla: {altura_pantalla}\nAnchura de pantalla: {anchura_pantalla}")
        x = (anchura_pantalla // 5) - (anchura//4)
        y = (altura_pantalla//5) - (altura//4)
        self.ventana.geometry(f"+{x}+{y}")
        self.ventana.title("Food OK!")
        self.ventana.config(bg="orange") 
        self.ventana.iconbitmap("C:\\FO_OK\\ico.ico")
        archivo_xlsx = 'CONTEO.xlsx'
        self.cont =0

        workbook = load_workbook('CONTEO.xlsx')
        sheet = workbook.active  
        self.items2 = []
        workbook.close()

        # Comprobar si el archivo existe
        if not os.path.exists(archivo_xlsx):
            # Si no existe, crear un nuevo archivo Excel
            self.book = Workbook()
            self.book.save(archivo_xlsx)
        # Cargar el archivo Excel existente
        self.book = load_workbook(archivo_xlsx)
        self.sheet = self.book.active
        #HACER COPIA PARA QUE NO SE BORRE
        self.elemntos_libro()
        self.book.save('CONTEO.xlsx')
        
        
        self.btns = {}
        self.datos1 = []
        self.SeleccionCompletos()
        self.SeleccionSanwich()
        self.operaciones()
        self.opciones()
        self.ventana.mainloop()
    #listo
    def opciones(self):

        self.img2 = ImageTk.PhotoImage(Image.open("C:\\FO_OK\\FOODOK.PNG").resize((180,160)))
        self.labelfoto = CTkLabel(self.ventana, text='', image = self.img2).place(x=950, y=20)
        self.botSan = ctk.CTkButton(self.ventana, text="SANDWICHES",width=120,height=30,border_width=0,corner_radius=20, command=lambda:self.mostrar_grupo1()).place(x=850, y=200)
        self.btnCompletos = ctk.CTkButton(self.ventana,text='COMPLETOS',width=120,height=30,border_width=0,corner_radius=20,command=lambda:self.mostrar_grupo2()).place(x=980, y=200)

        self.btnBebidas = CTkButton(self.ventana,text='MOSTRAR PRODUCTOS BD',width=120,height=30,border_width=0,corner_radius=20,).place(x=1110, y=200)

        self.btnPapasfritas = CTkButton(self.ventana,text='PAPASFRITAS',width=120,height=30,border_width=0,corner_radius=20,command=lambda:self.eliminar_pedido()).place(x=850, y=240)
        self.btnPollo = CTkButton(self.ventana,text='POLLO',width=120,height=30,border_width=0,corner_radius=20,).place(x=980, y=240)
        self.btnPizza = CTkButton(self.ventana,text='PIZZA',width=120,height=30,border_width=0,corner_radius=20,).place(x=1110, y=240)
        self.btnAgregados = CTkButton(self.ventana,text='AGREGADOS',width=120,height=30,border_width=0,corner_radius=20,).place(x=850, y=280)
        self.btnPichangas = CTkButton(self.ventana,text='PICHANGA',width=120,height=30,border_width=0,corner_radius=20,).place(x=980, y=280)
        self.btnColaciones = CTkButton(self.ventana,text='COLACIONES',width=120,height=30,border_width=0,corner_radius=20,).place(x=1110, y=280)
        self.btnPichangas = CTkButton(self.ventana,text='mostrar pedidos en la base de datos',width=120,height=30,border_width=0,corner_radius=20,command=lambda:self.mostrar_pedidos()).place(x=980, y=500)
        self.btnColaciones = CTkButton(self.ventana,text='eliminar pedido',width=120,height=30,border_width=0,corner_radius=20,command=lambda:self.eliminar_pedido()).place(x=1110, y=500)
    #listo
    def operaciones(self):
        self.img3 = ImageTk.PhotoImage(Image.open("C:\\FO_OK\\fi3.png").resize((25,25)))
        self.img4 = ImageTk.PhotoImage(Image.open("C:\\FO_OK\\faa.png").resize((25,25)))
        self.img5 = ImageTk.PhotoImage(Image.open("C:\\FO_OK\\nota.png").resize((25,25)))
        self.img6 = ImageTk.PhotoImage(Image.open("C:\\FO_OK\\x.png").resize((25,25)))
        self.img7 = ImageTk.PhotoImage(Image.open("C:\\FO_OK\\p.png").resize((25,25)))
        self.img8 = ImageTk.PhotoImage(Image.open("C:\\FO_OK\\r.png").resize((25,25)))
        self.img9 = ImageTk.PhotoImage(Image.open("C:\\FO_OK\\limpiar.png").resize((25,25)))
        self.img10 = ImageTk.PhotoImage(Image.open("C:\\FO_OK\\dividir.png").resize((25,25)))
        self.img11 = ImageTk.PhotoImage(Image.open("C:\\FO_OK\\anotar.png").resize((25,25)))
        self.img12 = ImageTk.PhotoImage(Image.open("C:\\FO_OK\\el.png").resize((25,25)))
        self.img13 = ImageTk.PhotoImage(Image.open("C:\\FO_OK\\fl1.png").resize((25,25)))
        self.img14 = ImageTk.PhotoImage(Image.open("C:\\FO_OK\\fl2.png").resize((25,25)))

        self.btnoper1 = CTkButton(self.ventana, text='Salida', width=120, height=30, border_color="black",fg_color="white", hover_color="gray90", text_color="black"
                              ,border_width=2, corner_radius=0, compound=ctk.TOP, image=self.img3).place(x=10, y=20)
        self.btnoper2 = CTkButton(self.ventana, text='Traslado', width=120, height=30, border_color="black",fg_color="white", hover_color="gray90", text_color="black"
                              ,border_width=2, corner_radius=0, compound=ctk.TOP, image=self.img4).place(x=130, y=20)
        self.btnoper3 = CTkButton(self.ventana, text='Nota Venta', width=120, height=30,border_color="black",fg_color="white", hover_color="gray90", text_color="black"
                             ,border_width=2, corner_radius=0, compound=ctk.TOP, image=self.img5).place(x=250, y=20)

        self.btnoper4 = CTkButton(self.ventana, text='Anular Pedido', width=120, height=30,border_color="black",fg_color="white", hover_color="gray90", text_color="black", 
                              border_width=2, corner_radius=0, compound=ctk.TOP, image=self.img6).place(x=10, y=78)
        self.btnoper5 = CTkButton(self.ventana, text='Pedientes', width=120, height=30, border_color="black",fg_color="white", hover_color="gray90", text_color="black",
                              border_width=2, corner_radius=0, compound=ctk.TOP, image=self.img7).place(x=130, y=78)
        self.btnoper6 = CTkButton(self.ventana, text='Reserva', width=120, height=30,border_color="black",fg_color="white", hover_color="gray90", text_color="black",
                                  border_width=2, corner_radius=0, compound=ctk.TOP, image=self.img8).place(x=250, y=78)
        
        self.btnoper7 = CTkButton(self.ventana, text='Lipiar Formulario', width=120, height=30,border_color="black",fg_color="white", hover_color="gray90", text_color="black", 
                              border_width=2, corner_radius=0, compound=ctk.TOP, image=self.img9).place(x=10, y=136)
        self.btnoper8 = CTkButton(self.ventana, text='Dividir', width=120, height=30, border_color="black",fg_color="white", hover_color="gray90", text_color="black",
                              border_width=2, corner_radius=0, compound=ctk.TOP, image=self.img10).place(x=130, y=136)
        self.btnoper9 = CTkButton(self.ventana, text='Anotación', width=120, height=30, border_color="black",fg_color="white", hover_color="gray90", text_color="black",
                              border_width=2, corner_radius=0, compound=ctk.TOP, image=self.img11).place(x=250, y=136)
        
        self.btnoper10 = CTkButton(self.ventana, text='Eliminar', width=120, height=30, border_color="black",fg_color="white", hover_color="gray90", text_color="black",
                              border_width=2, corner_radius=0, compound=ctk.TOP, image=self.img12).place(x=10, y=194)
        self.btnoper11 = CTkButton(self.ventana, text='Bajar', width=120, height=30, border_color="black",fg_color="white", hover_color="gray90", text_color="black",
                              border_width=2, corner_radius=0, compound=ctk.TOP, image=self.img13).place(x=130, y=194)
        self.btnoper12 = CTkButton(self.ventana, text='Subir', width=120, height=30, border_color="black",fg_color="white", hover_color="gray90", text_color="black",
                              border_width=2, corner_radius=0, compound=ctk.TOP, image=self.img14).place(x=250, y=194)
        
        self.btnoper10 = CTkButton(self.ventana, text='Eliminar', width=100, height=30, border_color="black",fg_color="white", hover_color="gray90", text_color="black",
                              border_width=2, corner_radius=0, compound=ctk.TOP, image=self.img12).place(x=410, y=20)
        self.btnoper11 = CTkButton(self.ventana, text='imprimir', width=100, height=30, border_color="black",fg_color="white", hover_color="gray90", text_color="black",
                              border_width=2, corner_radius=0, compound=ctk.TOP, image=self.img13).place(x=510, y=20)
        self.btnoper12 = CTkButton(self.ventana, text='Subir', width=100, height=30, border_color="black",fg_color="white", hover_color="gray90", text_color="black",
                              border_width=2, corner_radius=0, compound=ctk.TOP, image=self.img14).place(x=610, y=20)
        self.btnoper12 = CTkButton(self.ventana, text='Subir', width=100, height=30, border_color="black",fg_color="white", hover_color="gray90", text_color="black",
                              border_width=2, corner_radius=0, compound=ctk.TOP, image=self.img14).place(x=710, y=20)
        self.lista1 = CTkListbox(self.ventana, height=400,width=335)
        self.lista1.place(x=10,y=260)
    #listo
    def eliminar_pedido(self):
        id=CTkInputDialog(title='Eliminar producto', text='eliminar')
        id.geometry('300x200+800+400')
        dato=(int(id.get_input()))
        print(dato)
        contador_a_eliminar = [dato]  # Reemplaza con el contador del producto que deseas eliminar
        self.datos.eliminar_producto_por_contador(contador_a_eliminar)
        CTkMessagebox(title='MENSAJE', message=f'pedido {contador_a_eliminar} eliminado')
    #listo
    def mostrar_pedidos(self):
        print("hola")
        datos = self.datos.obtener_todos_los_productos()
        print(datos)
        display_string = "\n".join([" ".join(map(str, dato)) for dato in datos])

        CTkMessagebox(title='base_de_datos', message=display_string)
        #for dato in datos:
            #CTkMessagebox(title='base_de_datos', message=f'{dato}\n')
            
        pass






    #listo
    def insertar_elemento_en_excel(self, palabra, precio):
        self.contador += 1
        texto = self.contador
        
        
        self.lista1.insert(texto, palabra)
        # print("lo hace qui?")
        for row in self.sheet.iter_rows():
            for cell in row:
                if cell.value == "MECHADA" and palabra == "MECHADA":
                    self.cont1 += 1 
                    self.sheet[f'A{2}'] = "MECHADA"
                    self.sheet[f'B{2}'] = self.cont1

                   

                    contador_str = str("'" + str(self.contador) + "'")
                    categoria_str = 1

                    dato1 = self.datos.ingresar_producto(contador_str,categoria_str,palabra,precio,palabra,palabra)
                    

                if cell.value == "AVE" and palabra == "AVE":
                    self.cont2 += 1 
                    self.sheet[f'A{3}'] = "AVE"
                    self.sheet[f'B{3}'] = self.cont2

                    print(palabra,precio)

                    contador_str = str("'" + str(self.contador) + "'")
                    categoria_str = 1

                    dato1 = self.datos.ingresar_producto(contador_str,categoria_str,palabra,precio,palabra,palabra)

                if cell.value == "LOMO" and palabra == "LOMO":
                    self.cont3 += 1 
                    self.sheet[f'A{4}'] = "LOMO"
                    self.sheet[f'B{4}'] = self.cont3
                    print(palabra,precio)

                if cell.value == "CHURRASCO" and palabra == "CHURRASCO":
                    self.cont4 += 1 
                    self.sheet[f'A{5}'] = "CHURRASCO"
                    self.sheet[f'B{5}'] = self.cont4
                    print(palabra,precio)
                if cell.value == "MECHADA OK" and palabra == "MECHADA OK":
                    self.cont5 += 1 
                    self.sheet[f'A{6}'] = "MECHADA OK"
                    self.sheet[f'B{6}'] = self.cont5
                    print(palabra,precio)
                if cell.value == "MECHADA TRADICION" and palabra == "MECHADA TRADICION":
                    self.cont6 += 1 
                    self.sheet[f'A{7}'] = "MECHADA TRADICION"
                    self.sheet[f'B{7}'] = self.cont6
                
                if cell.value == "AVE OK" and palabra == "AVE OK":
                    self.cont7 += 1 
                    self.sheet[f'A{8}'] = "AVE OK"
                    self.sheet[f'B{8}'] = self.cont7
                
                if cell.value == "Completo Italiano" and palabra == "Completo Italiano":
                    self.cont8 += 1 
                    self.sheet[f'A{9}'] = "Completo Italiano"
                    self.sheet[f'B{9}'] = self.cont8
                
        self.book.save('CONTEO.xlsx')
        

        


        print(texto)
    #listo
    def elemntos_libro(self):
        workbook = load_workbook('CONTEO.xlsx')
        sheet = workbook.active  
        items = []

        for i, (item, value) in enumerate(items, start=2):
            self.sheet[f'A{i}'] = item
            self.sheet[f'B{i}'] = value

        workbook.close()
    # Llama a la función crear_boton para crear un botón para cada elemento en 'items'
    #listo
    def crear_boton_en_pantalla(self, elemento, precio, x, y):
    # Crear un botón con el elemento y el precio obtenidos
            btn = CTkButton(self.grupo1, text=elemento, width=180, height=30, border_width=0, corner_radius=20, command=lambda:self.insertar_elemento_en_excel(elemento, precio))
            btn.place(x=x, y=y)
    #listo
    
    def crear_botones(self, items2):
        
        workbook = load_workbook('PRECIO.xlsx')
        sheet = workbook.active  
        items2 = []
        print(items2)
        for row in sheet.iter_rows(min_row=2, max_col=2, values_only=True):
            elemento, precio = row
            items2.append((elemento, precio))
        workbook.close()
        
        for i, (elemento, precio) in enumerate(items2, start=2):
            print("boton creado")
            x = 10  # Posición x del botón
            y = 20 + (i - 2) * 40  # Ajusta la posición y para cada botón
            print(y)
            self.crear_boton_en_pantalla(elemento, precio, x, y) 



 # ver
    def cargar_datos_desde_excel(self, archivo):
        try:
            workbook = load_workbook(archivo)
            sheet = workbook.active
            datos = []
            for row in sheet.iter_rows(min_row=2, max_col=2, values_only=True):
                elemento, precio = row
                datos.append((elemento, precio))
            workbook.close()
            return datos
        except Exception as e:
            print(f"Error al cargar datos desde Excel: {e}")
            return []

    def Crear_Boleta(self, producto, precio):
        boleta = f"Producto: {producto}\nPrecio: {precio}"
        return boleta

    def generar_boletas_desde_excel(self, archivo):
        datos = self.cargar_datos_desde_excel(archivo)
        for producto, precio in datos:
            boleta = self.Crear_Boleta(producto, precio)
            print(boleta)

    def main(self, palabra, precio):
        # archivo_excel = 'precios.xlsx'
        print("entra_")
        print(palabra,precio)
        # self.generar_boletas_desde_excel(archivo_excel)
        workbook2 = load_workbook('BOLETA.xlsx')
        self.sheet = workbook2.active  
        self.contador2 += 1 
        self.cont += 1
        self.sheet[f'F{self.contador + 4}'] = self.cont
        self.sheet[f'H{self.contador + 4}'] = palabra
        self.sheet[f'I{self.contador + 4}'] = precio
        self.book.save('BOLETA.xlsx')
        self.book.close()
        pass

    #listo 
    def SeleccionSanwich(self):
        self.cont1 = 0
        self.cont2 = 0
        self.cont3 = 0
        self.cont4 = 0
        self.cont5 = 0
        self.cont6 = 0
        self.cont7 = 0
        self.cont8 = 0
        self.cont9 = 0
        self.contador = 0
        self.precio= 0  
        self.contador2 = 0
        self.cont = 0
    #listo 
    def SeleccionCompletos(self):
        self.btnAve = CTkButton(self.grupo2,text='Completo Italiano',width=180,height=30,border_width=0,corner_radius=20, command=lambda:(self.insertar_elemento("Completo Italiano"), self.insertar_precios(2500))).place(x=10, y=20)
        self.btnLomo = CTkButton(self.grupo2,text='Completo Italiano Chocrut',width=180,height=30,border_width=0,corner_radius=20,command=lambda:(self.insertar_elemento("Completo Italiano Chocrut"), self.insertar_precios(1000))).place(x=10, y=60)
        self.btnChurras = CTkButton(self.grupo2,text='Completo Chocrut',width=180,height=30,border_width=0,corner_radius=20,command=lambda:(self.insertar_elemento("Completo Chocrut"), self.insertar_precios(1000))).place(x=10, y=100)
     
        self.btnMechOK = CTkButton(self.grupo2,text='Completo Alemán',width=180,height=30,border_width=0,corner_radius=20,command=lambda:(self.insertar_elemento("Completo Alemán"), self.insertar_precios(1000))).place(x=10, y=140)
        self.btnMechTRA = CTkButton(self.grupo2,text='Completo Tomate-Mayo',width=180,height=30,border_width=0,corner_radius=20,command=lambda:(self.insertar_elemento("Completo Tomate-Mayo"), self.insertar_precios(1000))).place(x=10, y=180)
        
        self.btnAveOK = CTkButton(self.grupo2,text='Completo Palta-Mayo',width=180,height=30,border_width=0,corner_radius=20,command=lambda:(self.insertar_elemento("Completo Palta-Mayo"), self.insertar_precios(1000))).place(x=10, y=220)
        self.btnAveTRA = CTkButton(self.grupo2,text='Completo Mechada',width=180,height=30,border_width=0,corner_radius=20,command=lambda:(self.insertar_elemento("Completo Mechada"), self.insertar_precios(1000))).place(x=10, y=260)


    #listo
    def mostrar_grupo1(self):
        self.crear_botones(self.items2)
        self.grupo1.place(x=390,y=80)
        self.grupo2.place_forget()
        
    #listo
    def mostrar_grupo2(self):
        self.grupo1.place_forget()
        self.grupo2.place(x=390,y=80)
        
    
    
f =Interfaz()




















